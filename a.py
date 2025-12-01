#!/usr/bin/env python3
"""
Simple XLSX to CSV converter with CLI and main function.

Usage examples:
  python a.py input.xlsx output.csv
  python a.py input.xlsx output_dir --all-sheets
  python a.py input.xlsx output.csv --sheet Sheet1 --delimiter '\t'

Functions:
  convert_xlsx_to_csv(input_path, output_path, ...)

"""
from __future__ import annotations

import argparse
import csv
import os
import sys
from typing import Optional

from openpyxl import load_workbook


def _coerce_value(v):
	if v is None:
		return ""
	# For basic types, str provides a reasonable representation. For datetimes, isoformat is fine.
	try:
		return str(v)
	except Exception:
		return ""


def convert_xlsx_to_csv(
	input_path: str,
	output_path: str,
	sheet: Optional[str | int] = None,
	delimiter: str = ",",
	encoding: str = "utf-8",
	include_headers: bool = True,
	skip_blank_rows: bool = False,
	all_sheets: bool = False,
	overwrite: bool = False,
):
	"""
	Convert XLSX file to CSV.

	Parameters:
	  input_path: Path to .xlsx file
	  output_path: Path to .csv file or directory (when all_sheets=True)
	  sheet: sheet name or 0-based sheet index to convert (default first sheet)
	  delimiter: CSV delimiter string
	  encoding: Output file encoding
	  include_headers: whether to include header row if present
	  skip_blank_rows: skip blank rows from CSV output
	  all_sheets: convert all sheets, writing one CSV per sheet into output_path directory
	  overwrite: overwrite existing output files

	Returns: list of written output file paths
	"""
	if not os.path.exists(input_path):
		raise FileNotFoundError(f"Input file not found: {input_path}")

	wb = load_workbook(filename=input_path, read_only=True, data_only=True)
	outputs = []

	def write_sheet(ws, out_path):
		if os.path.exists(out_path) and not overwrite:
			raise FileExistsError(f"Output file {out_path} exists (use overwrite=True)")
		with open(out_path, "w", newline="", encoding=encoding) as f:
			writer = csv.writer(f, delimiter=delimiter)
			rows_written = 0
			for row in ws.iter_rows(values_only=True):
				if skip_blank_rows and all(v is None for v in row):
					continue
				writer.writerow([_coerce_value(v) for v in row])
				rows_written += 1
		return out_path

	if all_sheets:
		if not os.path.isdir(output_path):
			raise NotADirectoryError(f"When --all-sheets is used, output path must be a directory: {output_path}")
		for ws in wb.worksheets:
			safe_name = ws.title.replace(os.sep, "_")
			out_file = os.path.join(output_path, f"{safe_name}.csv")
			outputs.append(write_sheet(ws, out_file))
		return outputs

	# single sheet modality
	if sheet is None:
		ws = wb[wb.sheetnames[0]]
	else:
		if isinstance(sheet, int):
			ws = wb[wb.sheetnames[sheet]]
		else:
			ws = wb[sheet]

	outputs.append(write_sheet(ws, output_path))
	return outputs


def parse_args(argv=None):
	p = argparse.ArgumentParser(description="Convert .xlsx to .csv")
	p.add_argument("input", help="Path to input .xlsx file")
	p.add_argument("output", help="Path to output .csv file or dir (if --all-sheets)")
	p.add_argument("--sheet", help="Sheet name or 0-based index (defaults to first sheet)")
	p.add_argument("--delimiter", default=",", help="CSV delimiter (default: ,)")
	p.add_argument("--encoding", default="utf-8", help="Output encoding (default: utf-8)")
	p.add_argument("--all-sheets", action="store_true", help="Write all sheets as individual CSV files to output dir")
	p.add_argument("--overwrite", action="store_true", help="Overwrite existing output files")
	p.add_argument("--skip-blank-rows", action="store_true", help="Skip blank rows in output")
	return p.parse_args(argv)


def main(argv=None):
	args = parse_args(argv)
	sheet = None
	if args.sheet is not None:
		# try parse index
		try:
			sheet = int(args.sheet)
		except Exception:
			sheet = args.sheet

	try:
		output_files = convert_xlsx_to_csv(
			args.input,
			args.output,
			sheet=sheet,
			delimiter=args.delimiter,
			encoding=args.encoding,
			skip_blank_rows=args.skip_blank_rows,
			all_sheets=args.all_sheets,
			overwrite=args.overwrite,
		)
	except Exception as e:
		print(f"Error: {e}", file=sys.stderr)
		return 2

	for p in output_files:
		print(p)
	return 0


if __name__ == "__main__":
	raise SystemExit(main())

